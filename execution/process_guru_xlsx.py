"""
process_guru_xlsx.py
Processa os exports XLSX da Guru e salva em .tmp/guru_raw.csv

Filtros aplicados:
  - status == "Aprovada"
  - nome produto contém "Plataforma MEDsimple" (case-insensitive)
    → inclui "Plataforma MEDsimple" e "Plataforma MEDsimple (Vitalício)"
"""

import os
import csv
from datetime import datetime
import openpyxl

OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "..", ".tmp", "guru_raw.csv")

XLSX_FILES = [
    r"F:\Downloads\Guru-Vendas-2026-03-20-13-19-06.xlsx",
    r"F:\Downloads\Guru-Vendas-2026-03-20-13-18-54.xlsx",
    r"F:\Downloads\Guru-Vendas-2026-03-20-13-18-31.xlsx",
    r"F:\Downloads\Guru-Vendas-2026-03-20-13-18-15.xlsx",
    r"F:\Downloads\Guru-Vendas-2026-03-20-13-14-21.xlsx",
    r"F:\Downloads\Guru-Vendas-2026-03-20-13-10-47.xlsx",
]

PRODUTO_FILTRO = "plataforma medsimple"
STATUS_FILTRO = "aprovada"


def normalize_phone(code, number):
    digits = "".join(filter(str.isdigit, str(code or "") + str(number or "")))
    # Remover DDI +55 se presente (ex: 5511999... → 11999...)
    if digits.startswith("55") and len(digits) >= 12:
        digits = digits[2:]
    return digits


def parse_date(value):
    """Converte datetime do Excel ou string para ISO 8601."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    # Tentar string
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return None


def process_file(filepath):
    print(f"\n[Guru] Processando: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    # Ler cabeçalhos normalizando encoding corrompido
    raw_headers = [str(cell.value or "").strip() for cell in next(ws.iter_rows(max_row=1))]

    # Mapear colunas pelo conteúdo (encoding pode estar corrompido)
    col_map = {}
    for i, h in enumerate(raw_headers):
        h_lower = h.lower()
        if "email" in h_lower and "contato" in h_lower:
            col_map["email"] = i
        elif "nome" in h_lower and "contato" in h_lower and "empresa" not in h_lower:
            col_map["nome"] = i
        elif "nome" in h_lower and "produto" in h_lower:
            col_map["nome_produto"] = i
        elif "codigo" in h_lower and "telefone" in h_lower:
            col_map["codigo_tel"] = i
        elif "telefone" in h_lower and "contato" in h_lower:
            col_map["telefone"] = i
        elif "status" == h_lower:
            col_map["status"] = i
        elif "aprovacao" in h_lower or "aprovação" in h_lower:
            col_map["data_aprovacao"] = i

    print(f"  Colunas mapeadas: {col_map}")

    records = []
    skipped_status = 0
    skipped_produto = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        status = str(row[col_map["status"]] or "").strip().lower() if "status" in col_map else ""
        if status != STATUS_FILTRO:
            skipped_status += 1
            continue

        nome_produto = str(row[col_map.get("nome_produto", -1)] or "").lower() if "nome_produto" in col_map else ""
        if PRODUTO_FILTRO not in nome_produto:
            skipped_produto += 1
            continue

        email = str(row[col_map.get("email", -1)] or "").strip().lower() if "email" in col_map else ""
        nome = str(row[col_map.get("nome", -1)] or "").strip() if "nome" in col_map else ""
        code = row[col_map["codigo_tel"]] if "codigo_tel" in col_map else ""
        fone = row[col_map["telefone"]] if "telefone" in col_map else ""
        telefone = normalize_phone(code, fone)
        data = parse_date(row[col_map["data_aprovacao"]]) if "data_aprovacao" in col_map else None

        if not email and not telefone:
            continue
        if not data:
            continue

        records.append({
            "email": email,
            "telefone": telefone,
            "nome": nome,
            "data_compra": data,
            "plataforma": "guru",
        })

    wb.close()
    print(f"  Aprovadas + MEDsimple: {len(records)} | Ignoradas status: {skipped_status} | Produto errado: {skipped_produto}")
    return records


def main():
    print("=" * 55)
    print("GURU — Processando XLSXs históricos")
    print("=" * 55)

    all_records = []
    for f in XLSX_FILES:
        if not os.path.exists(f):
            print(f"[!] Arquivo não encontrado: {f}")
            continue
        all_records.extend(process_file(f))

    # Remover duplicatas exatas (mesma transação pode aparecer em múltiplos exports)
    seen = set()
    unique = []
    for r in all_records:
        key = (r["email"], r["telefone"], r["data_compra"])
        if key not in seen:
            seen.add(key)
            unique.append(r)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["email", "telefone", "nome", "data_compra", "plataforma"])
        writer.writeheader()
        writer.writerows(unique)

    print(f"\n[Guru] Total: {len(all_records)} | Únicos: {len(unique)}")
    print(f"[Guru] Salvo em: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
